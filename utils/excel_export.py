"""
Builds a complete, presentation-ready Excel workbook from a dataset:
Raw Data, Clean Data, Summary, Pivot Tables, a Dashboard sheet with native
Excel charts + KPI cards, and an "AI Search" reference sheet.

Design choices (see /mnt/skills/public/xlsx/SKILL.md conventions):
- Consistent professional font (Calibri) throughout
- Header rows are bold with a fill color; freeze panes + autofilter on data sheets
- Currency-shaped numeric columns get a "#,##0.00" number format
- No formulas that could error (#REF!/#DIV0!) - pivot/summary values are
  computed in pandas and written as static values, which is the safe choice
  when we don't control what messy data the user uploads.
"""
from __future__ import annotations

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from utils.charts import KPI_KEYWORDS, _is_identifier_column

NAVY = "1B2540"
TEAL = "22D3B8"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F4F8"

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(name="Calibri", bold=True, size=15, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="666666")
KPI_LABEL_FONT = Font(name="Calibri", size=10, color="666666")
KPI_VALUE_FONT = Font(name="Calibri", bold=True, size=18, color=NAVY)
BODY_FONT = Font(name="Calibri", size=10)
LINK_FONT = Font(name="Calibri", size=11, color="4F8EF7", underline="single")
THIN_BORDER = Border(bottom=Side(style="thin", color="DDDDDD"))

SHEET_ORDER = ["Navigation", "Dashboard", "Business Insights", "Summary", "Pivot Tables",
               "Clean Data", "Raw Data", "AI Search"]


def build_workbook(raw_df: pd.DataFrame, clean_df: pd.DataFrame, profile: dict,
                    cleaning_log: list[str], filename: str,
                    ai_qa_pairs: list[dict] | None = None,
                    business_insights: dict | None = None) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    _write_navigation_sheet(wb, filename, profile)
    _write_dashboard_sheet(wb, clean_df, profile, filename)
    _write_business_insights_sheet(wb, business_insights or {})
    _write_summary_sheet(wb, profile, cleaning_log)
    _write_pivot_sheet(wb, clean_df)
    _write_data_sheet(wb, "Clean Data", clean_df, apply_conditional_formatting=True)
    _write_data_sheet(wb, "Raw Data", raw_df)
    _write_ai_search_sheet(wb, clean_df, ai_qa_pairs or [])

    wb._sheets.sort(key=lambda s: SHEET_ORDER.index(s.title) if s.title in SHEET_ORDER else 99)
    wb.active = 0
    return wb


# --------------------------------------------------------------------------
# Navigation sheet
# --------------------------------------------------------------------------

def _write_navigation_sheet(wb: Workbook, filename: str, profile: dict):
    ws = wb.create_sheet("Navigation")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 55

    ws["B2"] = filename
    ws["B2"].font = Font(name="Calibri", bold=True, size=20, color=NAVY)
    ws["B3"] = f"Business Intelligence Workbook — {profile['shape']['rows']:,} rows, {profile['shape']['columns']} columns"
    ws["B3"].font = SUBTITLE_FONT

    entries = [
        ("Dashboard", "KPI cards and charts at a glance"),
        ("Business Insights", "Top/bottom performers, risks, and growth opportunities"),
        ("Summary", "Full column profile and cleaning log"),
        ("Pivot Tables", "Cross-tabulated metrics by category, with pivot charts"),
        ("Clean Data", "The cleaned dataset, ready to analyze"),
        ("Raw Data", "The original, unmodified upload"),
        ("AI Search", "Pre-answered business questions about this dataset"),
    ]
    row = 6
    for sheet_name, desc in entries:
        cell = ws.cell(row=row, column=2, value=f"→  {sheet_name}")
        cell.font = LINK_FONT
        cell.hyperlink = f"#'{sheet_name}'!A1"
        ws.cell(row=row, column=3, value=desc).font = BODY_FONT
        row += 2


# --------------------------------------------------------------------------
# Business Insights sheet
# --------------------------------------------------------------------------

def _write_business_insights_sheet(wb: Workbook, insights: dict):
    ws = wb.create_sheet("Business Insights")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 55

    row = 1
    _title(ws, f"A{row}", "Business Insights")
    row += 2

    def write_performer_table(title, items):
        nonlocal row
        if not items:
            return
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12, color=NAVY)
        row += 1
        _write_header_row(ws, row, ["Segment", "Metric", "Value", "Share %"])
        row += 1
        for item in items:
            ws.cell(row=row, column=1, value=item["label"])
            ws.cell(row=row, column=2, value=item["metric"])
            ws.cell(row=row, column=3, value=item["value"])
            ws.cell(row=row, column=4, value=f'{item["share_pct"]}%')
            row += 1
        row += 2

    write_performer_table("Top Performers", insights.get("top_performers", []))
    write_performer_table("Bottom Performers", insights.get("bottom_performers", []))

    growth = insights.get("growth_opportunities", [])
    if growth:
        ws.cell(row=row, column=1, value="Growth Opportunities").font = Font(bold=True, size=12, color=NAVY)
        row += 1
        for g in growth:
            ws.cell(row=row, column=1, value=f"• {g['label']}").font = Font(bold=True, size=10)
            ws.cell(row=row, column=2, value=g["reason"]).font = BODY_FONT
            ws.merge_cells(f"B{row}:C{row}")
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            row += 1
        row += 2

    risks = insights.get("risk_indicators", [])
    if risks:
        ws.cell(row=row, column=1, value="Risk Indicators").font = Font(bold=True, size=12, color="E05C5C")
        row += 1
        for r in risks:
            ws.cell(row=row, column=1, value=f"⚠ {r['message']}").font = BODY_FONT
            ws.merge_cells(f"A{row}:C{row}")
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1

    if not any([insights.get("top_performers"), insights.get("risk_indicators")]):
        ws.cell(row=row, column=1, value="Not enough categorical + numeric columns in this dataset "
                                          "to compute automatic business insights.").font = BODY_FONT


# --------------------------------------------------------------------------
# Dashboard sheet: KPI cards + native charts
# --------------------------------------------------------------------------

def _write_dashboard_sheet(wb: Workbook, df: pd.DataFrame, profile: dict, filename: str):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 14

    ws["B2"] = f"{filename} — Business Dashboard"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = f"{profile['shape']['rows']:,} rows · {profile['shape']['columns']} columns · Data quality score: {profile['quality_score']}/100"
    ws["B3"].font = SUBTITLE_FONT

    # ---- KPI cards ----
    numeric_cols = [c for c in df.select_dtypes(include=np.number).columns
                     if not _is_identifier_column(df, c)]
    metric_cols = [c for c in numeric_cols if any(k in c.lower() for k in KPI_KEYWORDS)][:4] or numeric_cols[:4]

    kpi_row = 5
    col_positions = ["B", "D", "F", "H"]
    for i, col in enumerate(metric_cols[:4]):
        series = df[col].dropna()
        if series.empty:
            continue
        pos = col_positions[i]
        label_cell = ws[f"{pos}{kpi_row}"]
        value_cell = ws[f"{pos}{kpi_row + 1}"]
        label_cell.value = f"Total {col.replace('_', ' ').title()}"
        label_cell.font = KPI_LABEL_FONT
        value_cell.value = round(float(series.sum()), 2)
        value_cell.font = KPI_VALUE_FONT
        value_cell.number_format = "#,##0.00"
        ws.merge_cells(f"{pos}{kpi_row}:{_shift(pos, 1)}{kpi_row}")
        for r in (kpi_row, kpi_row + 1):
            for c in (pos, _shift(pos, 1)):
                ws[f"{c}{r}"].fill = PatternFill("solid", fgColor=LIGHT_GREY)

    # ---- Data table backing the charts ----
    chart_data_row = kpi_row + 4
    ws[f"B{chart_data_row - 1}"] = "Chart Data (auto-generated)"
    ws[f"B{chart_data_row - 1}"].font = Font(name="Calibri", bold=True, size=10, color="999999")

    categorical_cols = [c for c in df.select_dtypes(include="object").columns
                         if 1 < df[c].nunique(dropna=True) <= 15]
    datetime_cols = df.select_dtypes(include="datetime").columns.tolist()

    next_row = chart_data_row
    chart_anchor_row = chart_data_row

    # Category breakdown chart
    if categorical_cols and metric_cols:
        cat_col, metric_col = categorical_cols[0], metric_cols[0]
        summary = df.groupby(cat_col, dropna=True)[metric_col].sum().sort_values(ascending=False).head(10)
        ws[f"B{next_row}"] = cat_col.replace("_", " ").title()
        ws[f"C{next_row}"] = f"Total {metric_col.replace('_', ' ').title()}"
        ws[f"B{next_row}"].font = ws[f"C{next_row}"].font = Font(bold=True, size=9)
        for i, (k, v) in enumerate(summary.items(), start=1):
            ws[f"B{next_row + i}"] = str(k)
            ws[f"C{next_row + i}"] = round(float(v), 2)
        chart = BarChart()
        chart.title = f"{metric_col.replace('_', ' ').title()} by {cat_col.replace('_', ' ').title()}"
        chart.y_axis.title = metric_col.replace("_", " ").title()
        data_ref = Reference(ws, min_col=3, min_row=next_row, max_row=next_row + len(summary))
        cats_ref = Reference(ws, min_col=2, min_row=next_row + 1, max_row=next_row + len(summary))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width, chart.height = 15, 8
        ws.add_chart(chart, f"B{chart_anchor_row + len(summary) + 3}")
        next_row += len(summary) + 20

    # Time series chart
    if datetime_cols and metric_cols:
        dcol, mcol = datetime_cols[0], metric_cols[0]
        ts = df[[dcol, mcol]].dropna()
        if not ts.empty:
            ts = ts.set_index(dcol)[mcol].resample("MS").sum()
            ws[f"F{chart_data_row}"] = "Month"
            ws[f"G{chart_data_row}"] = f"Total {mcol.replace('_', ' ').title()}"
            ws[f"F{chart_data_row}"].font = ws[f"G{chart_data_row}"].font = Font(bold=True, size=9)
            for i, (d, v) in enumerate(ts.items(), start=1):
                ws[f"F{chart_data_row + i}"] = d.strftime("%Y-%m")
                ws[f"G{chart_data_row + i}"] = round(float(v), 2)
            line = LineChart()
            line.title = f"{mcol.replace('_', ' ').title()} Over Time"
            data_ref = Reference(ws, min_col=7, min_row=chart_data_row, max_row=chart_data_row + len(ts))
            cats_ref = Reference(ws, min_col=6, min_row=chart_data_row + 1, max_row=chart_data_row + len(ts))
            line.add_data(data_ref, titles_from_data=True)
            line.set_categories(cats_ref)
            line.width, line.height = 15, 8
            ws.add_chart(line, f"F{chart_anchor_row + len(ts) + 3}")

    # Pie chart for the top categorical column's mix (row counts)
    if categorical_cols:
        cat_col = categorical_cols[0]
        counts = df[cat_col].value_counts(dropna=True).head(8)
        start = chart_data_row
        ws[f"J{start}"] = cat_col.replace("_", " ").title()
        ws[f"K{start}"] = "Count"
        ws[f"J{start}"].font = ws[f"K{start}"].font = Font(bold=True, size=9)
        for i, (k, v) in enumerate(counts.items(), start=1):
            ws[f"J{start + i}"] = str(k)
            ws[f"K{start + i}"] = int(v)
        pie = PieChart()
        pie.title = f"{cat_col.replace('_', ' ').title()} Mix"
        data_ref = Reference(ws, min_col=11, min_row=start, max_row=start + len(counts))
        cats_ref = Reference(ws, min_col=10, min_row=start + 1, max_row=start + len(counts))
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        pie.width, pie.height = 12, 8
        ws.add_chart(pie, f"J{chart_anchor_row + len(counts) + 3}")


def _shift(col_letter: str, n: int) -> str:
    idx = ord(col_letter) - ord("A")
    return chr(ord("A") + idx + n)


# --------------------------------------------------------------------------
# Summary sheet
# --------------------------------------------------------------------------

def _write_summary_sheet(wb: Workbook, profile: dict, cleaning_log: list[str]):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 45

    _title(ws, "A1", "Dataset Summary")
    ws["A3"] = "Rows"; ws["B3"] = profile["shape"]["rows"]
    ws["A4"] = "Columns"; ws["B4"] = profile["shape"]["columns"]
    ws["A5"] = "Memory (MB)"; ws["B5"] = profile["memory_usage_mb"]
    ws["A6"] = "Data Quality Score"; ws["B6"] = f"{profile['quality_score']} / 100"
    ws["A7"] = "Duplicate Rows Found"; ws["B7"] = profile["duplicate_rows"]
    for r in range(3, 8):
        ws[f"A{r}"].font = Font(bold=True, size=10)

    row = 10
    _title(ws, f"A{row}", "Column Profile")
    row += 1
    headers = ["Column", "Type", "Missing %", "Unique Values", "Cardinality %"]
    _write_header_row(ws, row, headers)
    row += 1
    for c in profile["columns"]:
        ws.cell(row=row, column=1, value=c["name"])
        ws.cell(row=row, column=2, value=c["dtype"])
        ws.cell(row=row, column=3, value=c["missing_pct"])
        ws.cell(row=row, column=4, value=c["unique_count"])
        ws.cell(row=row, column=5, value=c["cardinality_pct"])
        row += 1

    row += 2
    if profile.get("numeric_summary"):
        _title(ws, f"A{row}", "Numeric Summary")
        row += 1
        _write_header_row(ws, row, ["Column", "Mean", "Median", "Std", "Min", "P10", "Q1", "Q3", "P90", "Max", "Skew", "CV"])
        row += 1
        for col, s in profile["numeric_summary"].items():
            ws.cell(row=row, column=1, value=col)
            ws.cell(row=row, column=2, value=s["mean"])
            ws.cell(row=row, column=3, value=s["median"])
            ws.cell(row=row, column=4, value=s["std"])
            ws.cell(row=row, column=5, value=s["min"])
            ws.cell(row=row, column=6, value=s.get("p10"))
            ws.cell(row=row, column=7, value=s["q1"])
            ws.cell(row=row, column=8, value=s["q3"])
            ws.cell(row=row, column=9, value=s.get("p90"))
            ws.cell(row=row, column=10, value=s["max"])
            ws.cell(row=row, column=11, value=s["skewness"])
            ws.cell(row=row, column=12, value=s.get("coefficient_of_variation"))
            row += 1
        row += 2

    _title(ws, f"A{row}", "Cleaning Steps Applied")
    row += 1
    for step in cleaning_log:
        ws.cell(row=row, column=1, value=f"• {step}").font = BODY_FONT
        ws.merge_cells(f"A{row}:F{row}")
        row += 1


def _title(ws: Worksheet, cell: str, text: str):
    ws[cell] = text
    ws[cell].font = TITLE_FONT


def _write_header_row(ws: Worksheet, row: int, headers: list[str]):
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")


# --------------------------------------------------------------------------
# Pivot sheet
# --------------------------------------------------------------------------

def _write_pivot_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Pivot Tables")
    ws.column_dimensions["A"].width = 22

    numeric_cols = [c for c in df.select_dtypes(include=np.number).columns
                     if not _is_identifier_column(df, c)][:2]
    categorical_cols = [c for c in df.select_dtypes(include="object").columns
                         if 1 < df[c].nunique(dropna=True) <= 30][:3]

    if not numeric_cols or not categorical_cols:
        ws["A1"] = "Not enough categorical + numeric columns to build pivot tables for this dataset."
        return

    row = 1
    first_table_range = None
    for t_idx, cat_col in enumerate(categorical_cols):
        for m_idx, metric_col in enumerate(numeric_cols):
            pivot = df.pivot_table(index=cat_col, values=metric_col, aggfunc=["sum", "mean", "count"])
            pivot.columns = ["Sum", "Average", "Count"]
            pivot = pivot.sort_values("Sum", ascending=False)

            _title(ws, f"A{row}", f"{metric_col.replace('_', ' ').title()} by {cat_col.replace('_', ' ').title()}")
            row += 1
            table_start_row = row
            _write_header_row(ws, row, [cat_col.title()] + list(pivot.columns))
            row += 1
            for idx, vals in pivot.iterrows():
                ws.cell(row=row, column=1, value=str(idx))
                for j, v in enumerate(vals, start=2):
                    ws.cell(row=row, column=j, value=round(float(v), 2))
                row += 1
            table_end_row = row - 1

            if t_idx == 0 and m_idx == 0:
                first_table_range = (table_start_row, table_end_row)
            row += 2

    # Pivot chart for the first table, placed to the right so it doesn't
    # overlap any of the stacked pivot tables below it.
    if first_table_range:
        start, end = first_table_range
        chart = BarChart()
        chart.title = "Pivot Chart — First Table (Sum)"
        data_ref = Reference(ws, min_col=2, min_row=start, max_row=end)
        cats_ref = Reference(ws, min_col=1, min_row=start + 1, max_row=end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width, chart.height = 16, 9
        ws.add_chart(chart, f"F{start}")


# --------------------------------------------------------------------------
# Raw / Clean data sheets
# --------------------------------------------------------------------------

def _write_data_sheet(wb: Workbook, name: str, df: pd.DataFrame, apply_conditional_formatting: bool = False):
    ws = wb.create_sheet(name)
    # Cap at a comfortable size for a generated file; full data is always
    # available via the CSV export endpoints too.
    display_df = df.head(20000)

    _write_header_row(ws, 1, list(display_df.columns))
    numeric_cols = set(display_df.select_dtypes(include=np.number).columns)

    for r_idx, row in enumerate(display_df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=_excel_safe(value))
            cell.font = BODY_FONT
            if display_df.columns[c_idx - 1] in numeric_cols:
                cell.number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    if len(display_df):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(display_df.columns))}{len(display_df) + 1}"
    for i, col in enumerate(display_df.columns, start=1):
        width = min(28, max(10, int(display_df[col].astype(str).str.len().head(200).mean() or 10) + 4))
        ws.column_dimensions[get_column_letter(i)].width = width

    if apply_conditional_formatting and len(display_df):
        last_row = len(display_df) + 1
        for i, col in enumerate(display_df.columns, start=1):
            if col not in numeric_cols:
                continue
            col_letter = get_column_letter(i)
            rng = f"{col_letter}2:{col_letter}{last_row}"
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="min", start_color="F7C6C6",
                mid_type="percentile", mid_value=50, mid_color="FFF6D6",
                end_type="max", end_color="C6F0DE",
            ))


def _excel_safe(value):
    if pd.isna(value):
        return None
    if isinstance(value, (pd.Timestamp,)):
        return value.to_pydatetime()
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    return value


# --------------------------------------------------------------------------
# AI Search sheet
# --------------------------------------------------------------------------

def _write_ai_search_sheet(wb: Workbook, df: pd.DataFrame, ai_qa_pairs: list[dict]):
    ws = wb.create_sheet("AI Search")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 70

    _title(ws, "A1", "AI Search — Answers Generated at Export Time")
    ws["A2"] = (
        "This sheet is a snapshot: each question below was answered by the AI "
        "analyst using the dataset at the moment this workbook was exported. "
        "Re-export from Datagrid to refresh answers, or use the live AI Chat "
        "in the app for open-ended questions."
    )
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:B2")
    ws.row_dimensions[2].height = 34

    row = 4
    _write_header_row(ws, row, ["Question", "Answer"])
    row += 1

    pairs = ai_qa_pairs or _fallback_qa_pairs(df)
    for qa in pairs:
        ws.cell(row=row, column=1, value=qa["question"]).font = Font(bold=True, size=10)
        answer_cell = ws.cell(row=row, column=2, value=qa["answer"])
        answer_cell.font = BODY_FONT
        answer_cell.alignment = Alignment(wrap_text=True)
        row += 1


def _fallback_qa_pairs(df: pd.DataFrame) -> list[dict]:
    """Used when AI is disabled: deterministic pandas-computed Q&A so the
    sheet is never empty."""
    pairs = []
    numeric_cols = [c for c in df.select_dtypes(include=np.number).columns
                     if not _is_identifier_column(df, c)]
    categorical_cols = [c for c in df.select_dtypes(include="object").columns
                         if 1 < df[c].nunique(dropna=True) <= 30]

    if numeric_cols:
        col = numeric_cols[0]
        pairs.append({
            "question": f"What is the total {col.replace('_', ' ')}?",
            "answer": f"{df[col].sum():,.2f}",
        })
        pairs.append({
            "question": f"What is the average {col.replace('_', ' ')}?",
            "answer": f"{df[col].mean():,.2f}",
        })
        if not df[col].isna().all():
            top_row = df.loc[df[col].idxmax()]
            pairs.append({
                "question": f"Which row has the highest {col.replace('_', ' ')}?",
                "answer": ", ".join(f"{k}: {v}" for k, v in top_row.head(6).items()),
            })

    if categorical_cols and numeric_cols:
        cat, metric = categorical_cols[0], numeric_cols[0]
        best = df.groupby(cat)[metric].sum().idxmax()
        pairs.append({
            "question": f"Which {cat.replace('_', ' ')} has the highest total {metric.replace('_', ' ')}?",
            "answer": str(best),
        })

    if categorical_cols:
        cat = categorical_cols[0]
        top = df[cat].value_counts().idxmax()
        pairs.append({
            "question": f"What is the most common {cat.replace('_', ' ')}?",
            "answer": f"{top} ({int(df[cat].value_counts().max())} occurrences)",
        })

    pairs.append({
        "question": "How many rows and columns does this dataset have?",
        "answer": f"{df.shape[0]:,} rows and {df.shape[1]} columns.",
    })
    return pairs
